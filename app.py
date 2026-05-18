import streamlit as st
import pandas as pd
from datetime import datetime, date
import pytz
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from supabase import create_client, Client
from pyzbar.pyzbar import decode
from PIL import Image, ImageEnhance, ImageFilter
import numpy as np
import plotly.express as px
import io

try:
    import qrcode
    QR_OK = True
except ImportError:
    QR_OK = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    RL_OK = True
except ImportError:
    RL_OK = False

st.set_page_config(
    page_title="CCB Musical — Check-in",
    page_icon="🎵",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# =============================================================
# PALETA CCB OFICIAL
# Fonte: Manual de Identidade Visual CCB (2024)
# Azul escuro: #1C3D5A  |  Verde-azulado: #49656C
# Verde menta: #5DB196  |  Verde claro: #98CDBD
# Cinza:       #A5A5A5  |  Branco: #F8F9FA
# =============================================================
CSS = """
<style>
/* ===== RESET ===== */
[data-testid="stSidebar"],
[data-testid="collapsedControl"] { display: none !important; }
#MainMenu, footer, header         { visibility: hidden !important; }

/* ===== FUNDO ===== */
.stApp {
    background: #0e1c2a !important;
    font-family: 'Inter', 'Segoe UI', sans-serif !important;
}
.block-container {
    padding: 0 1.5rem 2rem !important;
    max-width: 1100px !important;
}

/* ===== TOPBAR ===== */
.ccb-topbar {
    background: linear-gradient(90deg, #0b1622 0%, #1C3D5A 100%);
    border-bottom: 2px solid #5DB196;
    padding: 14px 24px 12px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin: -1rem -1.5rem 0;
    position: sticky;
    top: 0;
    z-index: 999;
}
.ccb-topbar-logo {
    display: flex; align-items: center; gap: 10px;
}
.ccb-topbar-logo .logo-icon { font-size: 1.4rem; }
.ccb-topbar-logo .logo-text {
    font-size: 1rem; font-weight: 800;
    color: #fff; letter-spacing: 0.5px;
}
.ccb-topbar-logo .logo-sub {
    font-size: 0.7rem; color: #98CDBD;
    letter-spacing: 1.5px; text-transform: uppercase;
    margin-top: 1px;
}
.ccb-topbar-badge {
    background: rgba(93,177,150,0.15);
    border: 1px solid #5DB196;
    border-radius: 99px;
    padding: 4px 14px;
    color: #98CDBD;
    font-size: 0.72rem;
    font-weight: 600;
    letter-spacing: 1px;
    text-transform: uppercase;
}

/* ===== NAV PILLS ===== */
.ccb-nav {
    display: flex;
    gap: 6px;
    padding: 14px 0 4px;
    flex-wrap: wrap;
    border-bottom: 1px solid rgba(93,177,150,0.15);
    margin-bottom: 22px;
}
.ccb-nav-btn-wrap { flex: 1; min-width: 100px; }
.ccb-nav-pill {
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 6px;
    width: 100%;
    padding: 10px 8px;
    border-radius: 10px;
    font-size: 0.8rem;
    font-weight: 700;
    letter-spacing: 0.5px;
    text-transform: uppercase;
    color: #A5A5A5;
    background: rgba(255,255,255,0.03);
    border: 1px solid rgba(255,255,255,0.06);
    transition: all 0.18s;
    cursor: pointer;
    white-space: nowrap;
}
.ccb-nav-pill:hover {
    background: rgba(93,177,150,0.1);
    color: #98CDBD;
    border-color: rgba(93,177,150,0.3);
}
.ccb-nav-pill.active {
    background: linear-gradient(135deg, #1C3D5A, #49656C);
    color: #fff;
    border-color: #5DB196;
    box-shadow: 0 2px 12px rgba(93,177,150,0.25);
}
.ccb-nav-pill .pill-icon { font-size: 1rem; }

/* ===== PAGE HEADER ===== */
.page-header {
    display: flex; align-items: center; gap: 14px;
    padding: 18px 22px;
    background: linear-gradient(135deg, #132233 0%, #1C3D5A 60%, #2a5240 100%);
    border: 1px solid rgba(93,177,150,0.3);
    border-radius: 16px;
    margin-bottom: 20px;
    animation: fadeUp 0.3s ease;
}
.page-header .ph-icon { font-size: 2rem; }
.page-header .ph-title { color: #fff; font-size: 1.35rem; font-weight: 800; margin: 0; }
.page-header .ph-sub   { color: #98CDBD; font-size: 0.82rem; margin: 3px 0 0; }

/* ===== METRIC CARDS ===== */
.stat-row { display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 20px; }
.stat-card {
    flex: 1; min-width: 110px;
    background: rgba(28,61,90,0.35);
    border: 1px solid rgba(93,177,150,0.2);
    border-radius: 14px;
    padding: 16px 12px;
    text-align: center;
    transition: transform 0.18s;
}
.stat-card:hover { transform: translateY(-2px); }
.stat-val { font-size: 2rem; font-weight: 800; margin: 0; line-height: 1; }
.stat-lbl { font-size: 0.68rem; text-transform: uppercase; letter-spacing: 1.5px; margin-top: 5px; color: #A5A5A5; }
.sv-blue  { color: #60b4d4; }
.sv-green { color: #5DB196; }
.sv-teal  { color: #98CDBD; }
.sv-gray  { color: #A5A5A5; }

/* ===== PROGRESSO ===== */
.prog-bar-wrap {
    background: rgba(255,255,255,0.06);
    border-radius: 99px; height: 8px;
    margin: 0 0 20px; overflow: hidden;
}
.prog-bar-fill {
    height: 100%; border-radius: 99px;
    background: linear-gradient(90deg, #1C3D5A, #49656C, #5DB196);
    transition: width 0.6s ease;
}

/* ===== REUNIAO CARD ===== */
.reuniao-card {
    background: rgba(28,61,90,0.25);
    border: 1px solid rgba(93,177,150,0.2);
    border-left: 3px solid #5DB196;
    border-radius: 12px;
    padding: 16px 20px;
    margin: 8px 0;
    display: flex; align-items: center; justify-content: space-between;
    transition: all 0.18s;
}
.reuniao-card:hover { background: rgba(28,61,90,0.45); transform: translateX(3px); }
.reuniao-nome { color: #e2e8f0; font-size: 1rem; font-weight: 700; margin: 0; }
.reuniao-meta { color: #A5A5A5; font-size: 0.78rem; margin: 3px 0 0; }
.badge-hoje {
    background: #5DB196; color: #0e1c2a;
    font-size: 0.65rem; font-weight: 800;
    padding: 3px 10px; border-radius: 99px;
    text-transform: uppercase; letter-spacing: 1px;
}
.badge-fut {
    background: rgba(93,177,150,0.12);
    color: #98CDBD; border: 1px solid rgba(93,177,150,0.3);
    font-size: 0.65rem; font-weight: 700;
    padding: 3px 10px; border-radius: 99px;
}

/* ===== SECAO ===== */
.sec-label {
    color: #5DB196; font-size: 0.68rem; font-weight: 800;
    text-transform: uppercase; letter-spacing: 2px;
    margin: 18px 0 10px;
    display: flex; align-items: center; gap: 8px;
}
.sec-label::after { content:''; flex:1; height:1px; background: rgba(93,177,150,0.2); }

/* ===== FEEDBACK ===== */
.fb-ok   { background: linear-gradient(135deg,#0a2e20,#0d3d29); border:1px solid #5DB196; border-radius:14px; padding:20px 24px; text-align:center; margin:10px 0; animation: bounceIn .35s ease; }
.fb-ok .fb-title { color:#98CDBD; font-size:1rem; font-weight:600; }
.fb-ok .fb-nome  { color:#fff; font-size:1.4rem; font-weight:800; margin:4px 0; }
.fb-ok .fb-sub   { color:#5DB196; font-size:.85rem; }
.fb-warn { background:rgba(120,83,0,.3); border:1px solid #d4a017; border-radius:14px; padding:16px 20px; text-align:center; margin:10px 0; }
.fb-warn .fb-t   { color:#f5d070; font-size:.95rem; font-weight:700; }
.fb-erro { background:rgba(120,0,0,.3); border:1px solid #c0392b; border-radius:14px; padding:16px 20px; text-align:center; margin:10px 0; }
.fb-erro .fb-t   { color:#f1948a; font-size:.95rem; font-weight:700; }
.fb-idle { background:rgba(28,61,90,.2); border:1px dashed rgba(93,177,150,.3); border-radius:14px; padding:16px 20px; text-align:center; margin:10px 0; }
.fb-idle .fb-t   { color:#98CDBD; font-size:.88rem; }

/* ===== INPUTS ===== */
input, textarea,
[data-baseweb="input"] input {
    background: rgba(28,61,90,0.3) !important;
    border: 1px solid rgba(93,177,150,0.3) !important;
    border-radius: 9px !important;
    color: #e2e8f0 !important;
}
input:focus, textarea:focus {
    border-color: #5DB196 !important;
    box-shadow: 0 0 0 2px rgba(93,177,150,0.2) !important;
}
[data-baseweb="select"] > div {
    background: rgba(28,61,90,0.3) !important;
    border: 1px solid rgba(93,177,150,0.3) !important;
    border-radius: 9px !important;
    color: #e2e8f0 !important;
}

/* ===== BOTOES ===== */
.stButton > button {
    border-radius: 9px !important;
    font-weight: 600 !important;
    transition: all 0.18s !important;
    background: rgba(28,61,90,0.4) !important;
    border: 1px solid rgba(93,177,150,0.3) !important;
    color: #98CDBD !important;
}
.stButton > button:hover {
    background: rgba(73,101,108,0.5) !important;
    border-color: #5DB196 !important;
    transform: translateY(-1px) !important;
}
.stButton > button[kind="primary"],
.stButton > button[data-testid="baseButton-primary"] {
    background: linear-gradient(135deg,#1C3D5A,#49656C) !important;
    border: 1px solid #5DB196 !important;
    color: #fff !important;
    box-shadow: 0 3px 12px rgba(93,177,150,0.25) !important;
}
.stButton > button[kind="primary"]:hover {
    background: linear-gradient(135deg,#234d72,#5DB196) !important;
    box-shadow: 0 5px 18px rgba(93,177,150,0.4) !important;
}

/* ===== TABS INTERNAS ===== */
[data-testid="stTabs"] button[role="tab"] {
    background: rgba(28,61,90,0.25) !important;
    border-radius: 8px 8px 0 0 !important;
    color: #A5A5A5 !important;
    font-size: 0.82rem !important;
    font-weight: 600 !important;
    border: 1px solid rgba(93,177,150,0.12) !important;
    border-bottom: none !important;
    padding: 9px 16px !important;
    letter-spacing: 0.3px !important;
}
[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
    background: linear-gradient(135deg,#1C3D5A,#49656C) !important;
    color: #fff !important;
    border-color: #5DB196 !important;
}

/* ===== DATAFRAME ===== */
[data-testid="stDataFrame"] {
    border-radius: 10px !important;
    border: 1px solid rgba(93,177,150,0.2) !important;
    overflow: hidden !important;
}

/* ===== METRIC NATIVO ===== */
[data-testid="stMetric"] {
    background: rgba(28,61,90,0.3) !important;
    border: 1px solid rgba(93,177,150,0.2) !important;
    border-radius: 12px !important; padding: 14px !important;
}
[data-testid="stMetricValue"] { color: #5DB196 !important; }
[data-testid="stMetricLabel"] { color: #A5A5A5 !important; font-size: 0.75rem !important; }

/* ===== CAMERA ===== */
[data-testid="stCameraInput"] > div {
    border-radius: 14px !important;
    border: 2px solid rgba(93,177,150,0.4) !important;
    overflow: hidden !important;
}

/* ===== SCROLLBAR ===== */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: rgba(255,255,255,0.02); }
::-webkit-scrollbar-thumb { background: rgba(93,177,150,0.35); border-radius: 99px; }

/* ===== ANIMS ===== */
@keyframes fadeUp   { from { opacity:0; transform:translateY(12px); } to { opacity:1; transform:translateY(0); } }
@keyframes bounceIn { 0% { transform:scale(.9); opacity:0; } 60% { transform:scale(1.03); opacity:1; } 100% { transform:scale(1); } }

div[data-testid="stHorizontalBlock"] > div[data-testid="column"] button {
    min-height: 48px !important;
}
</style>
"""
st.markdown(CSS, unsafe_allow_html=True)

# -------------------------------------------------------
# SUPABASE
# -------------------------------------------------------
@st.cache_resource
def get_supabase():
    return create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])

supabase: Client = get_supabase()
BR = pytz.timezone("America/Cuiaba")

def agora_br():   return datetime.now(BR)

def load_participantes():
    r = supabase.table("participantes").select("*").order("nome").execute()
    return pd.DataFrame(r.data) if r.data else pd.DataFrame()

def load_reunioes():
    return supabase.table("reunioes").select("*").execute().data or []

def load_presencas(meeting_id=None):
    q = supabase.table("presencas").select("*, participantes(nome,cargo,localidade,instrumento)")
    if meeting_id: q = q.eq("meeting_id", meeting_id)
    return q.execute().data or []

def gerar_pdf_membros(membros: list) -> bytes:
    NAVY  = colors.HexColor("#1C3D5A")
    TEAL  = colors.HexColor("#5DB196")
    GRAY  = colors.HexColor("#555555")
    BLACK = colors.black
    LIGHT = colors.HexColor("#f0f5f3")
    WHITE = colors.white
    hs = ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=10, textColor=WHITE,  alignment=TA_CENTER)
    ns = ParagraphStyle("n", fontName="Helvetica-Bold", fontSize=11, textColor=BLACK,  alignment=TA_LEFT,   leading=14)
    ls = ParagraphStyle("l", fontName="Helvetica",      fontSize=9,  textColor=GRAY,   alignment=TA_LEFT)
    cs = ParagraphStyle("c", fontName="Helvetica-Bold", fontSize=13, textColor=NAVY,   alignment=TA_CENTER)
    buf_pdf = io.BytesIO()
    doc = SimpleDocTemplate(buf_pdf, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []
    for m in membros:
        qr = qrcode.QRCode(version=2, box_size=6, border=2)
        qr.add_data(m["id"]); qr.make(fit=True)
        qi = qr.make_image(fill_color="#1C3D5A", back_color="white")
        bq = io.BytesIO(); qi.save(bq, format="PNG"); bq.seek(0)
        ri = RLImage(bq, width=3.5*cm, height=3.5*cm)
        info = [Paragraph(m["nome"], ns), Spacer(1,4),
                Paragraph(f'<font color="#555">Cargo:</font> <b>{m["cargo"]}</b>', ls),
                Paragraph(f'<font color="#555">Local:</font> <b>{m.get("localidade","")}</b>', ls),
                Paragraph(f'<font color="#555">Instr.:</font> <b>{m.get("instrumento","")}</b>', ls)]
        card = Table([[Paragraph("QR CODE DE CHECK-IN — CCB MUSICAL", hs),""],
                      [ri, info],
                      [Paragraph(m["id"], cs),""]], colWidths=[4*cm,12*cm])
        card.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),NAVY),("SPAN",(0,0),(-1,0)),
            ("TOPPADDING",(0,0),(-1,0),6),("BOTTOMPADDING",(0,0),(-1,0),6),
            ("VALIGN",(0,1),(-1,1),"MIDDLE"),("ALIGN",(0,1),(0,1),"CENTER"),
            ("LEFTPADDING",(1,1),(1,1),10),("TOPPADDING",(0,1),(-1,1),8),("BOTTOMPADDING",(0,1),(-1,1),8),
            ("SPAN",(0,2),(-1,2)),("BACKGROUND",(0,2),(-1,2),LIGHT),
            ("TOPPADDING",(0,2),(-1,2),6),("BOTTOMPADDING",(0,2),(-1,2),6),
            ("BOX",(0,0),(-1,-1),1.2,TEAL),
        ]))
        story.append(card); story.append(Spacer(1,0.7*cm))
    doc.build(story); buf_pdf.seek(0)
    return buf_pdf.read()

# -------------------------------------------------------
# NAVEGACAO
# -------------------------------------------------------
PAGINAS = [
    ("home",    "🏠", "Painel"),
    ("checkin", "📷", "Check-in"),
    ("relat",   "📊", "Relatórios"),
    ("membros", "👥", "Membros"),
]

if "pagina" not in st.session_state:
    st.session_state.pagina = "home"

# ---- Top bar ----
st.markdown('''
<div class="ccb-topbar">
  <div class="ccb-topbar-logo">
    <span class="logo-icon">🎵</span>
    <div>
      <div class="logo-text">CCB Musical</div>
      <div class="logo-sub">Congregação Cristã no Brasil</div>
    </div>
  </div>
  <div class="ccb-topbar-badge">Sistema de Presença</div>
</div>
''', unsafe_allow_html=True)

st.markdown('<div style="height:10px"></div>', unsafe_allow_html=True)

# ---- Pills de navegacao (HTML visual + botoes reais sobrepostos) ----
ativo = st.session_state.pagina
pills_html = '<div class="ccb-nav">'
for key, icon, label in PAGINAS:
    cls = "ccb-nav-pill active" if ativo == key else "ccb-nav-pill"
    pills_html += f'<div class="ccb-nav-btn-wrap"><div class="{cls}"><span class="pill-icon">{icon}</span>{label}</div></div>'
pills_html += '</div>'
st.markdown(pills_html, unsafe_allow_html=True)

cols_nav = st.columns(4)
for i, (key, icon, label) in enumerate(PAGINAS):
    with cols_nav[i]:
        if st.button(f"{icon} {label}", key=f"nav_{key}", use_container_width=True):
            st.session_state.pagina = key
            st.rerun()

pagina = st.session_state.pagina

# ===================================================
# PAINEL HOME
# ===================================================
if pagina == "home":
    st.markdown('''
    <div class="page-header">
        <span class="ph-icon">🏠</span>
        <div><p class="ph-title">Painel Geral</p>
        <p class="ph-sub">Visão rápida do sistema</p></div>
    </div>''', unsafe_allow_html=True)

    df_p = load_participantes()
    reun = load_reunioes()
    pres = load_presencas()
    tm   = len(df_p); tr = len(reun); tp = len(pres)
    pct  = round(tp / tm * 100) if tm else 0

    st.markdown(f'''
    <div class="stat-row">
        <div class="stat-card"><p class="stat-val sv-blue">{tm}</p><p class="stat-lbl">Membros</p></div>
        <div class="stat-card"><p class="stat-val sv-teal">{tr}</p><p class="stat-lbl">Reuniões</p></div>
        <div class="stat-card"><p class="stat-val sv-green">{tp}</p><p class="stat-lbl">Presenças</p></div>
        <div class="stat-card"><p class="stat-val sv-gray">{pct}%</p><p class="stat-lbl">Frequência</p></div>
    </div>
    <div class="prog-bar-wrap"><div class="prog-bar-fill" style="width:{pct}%"></div></div>
    ''', unsafe_allow_html=True)

    st.markdown('<div class="sec-label">📅 Reuniões</div>', unsafe_allow_html=True)
    hoje = agora_br().date()
    if not reun:
        st.info("Nenhuma reunião cadastrada ainda.")
    for r in sorted(reun, key=lambda x: x["data"]):
        dr = date.fromisoformat(r["data"]) if isinstance(r["data"], str) else r["data"]
        if dr == hoje:
            badge = '<span class="badge-hoje">Hoje</span>'
        else:
            badge = f'<span class="badge-fut">{dr.strftime("%d/%m")}</span>'
        st.markdown(f'''
        <div class="reuniao-card">
            <div>
                <p class="reuniao-nome">{r["nome"]}</p>
                <p class="reuniao-meta">{r.get("horario","")} &nbsp;·&nbsp; {dr.strftime("%d/%m/%Y")}</p>
            </div>
            {badge}
        </div>''', unsafe_allow_html=True)

# ===================================================
# CHECK-IN
# ===================================================
elif pagina == "checkin":
    st.markdown('''
    <div class="page-header">
        <span class="ph-icon">📷</span>
        <div><p class="ph-title">Check-in</p>
        <p class="ph-sub">Escaneie ou digite o código do membro</p></div>
    </div>''', unsafe_allow_html=True)

    reun = load_reunioes()
    if not reun:
        st.warning("Nenhuma reunião cadastrada."); st.stop()

    sel    = st.selectbox("📅 Selecione a reunião", [r["nome"] for r in reun])
    reuniao = next(r for r in reun if r["nome"] == sel)

    tab_cam, tab_cod = st.tabs(["📸  Câmera QR", "⌨️  Código Manual"])

    def registrar(codigo, mid):
        codigo = codigo.upper().strip()
        res = supabase.table("participantes").select("*").eq("id", codigo).execute()
        if not res.data: return None, "nf"
        m = res.data[0]
        dup = supabase.table("presencas").select("id").eq("id_participante",codigo).eq("meeting_id",mid).execute()
        if dup.data: return m, "dup"
        supabase.table("presencas").insert({
            "id_participante":codigo, "meeting_id":mid,
            "nome":m["nome"], "cargo":m["cargo"], "localidade":m["localidade"],
            "horario":agora_br().strftime("%H:%M:%S"),
            "data_registro":agora_br().strftime("%Y-%m-%d")
        }).execute()
        return m, "ok"

    with tab_cam:
        img = st.camera_input("Aponte a câmera para o QR Code")
        if img:
            pil = Image.open(img).convert("L")
            pil = ImageEnhance.Contrast(pil).enhance(2.5).filter(ImageFilter.SHARPEN)
            dec = decode(np.array(pil))
            if dec:
                cod = dec[0].data.decode().strip()
                m, s = registrar(cod, reuniao["id"])
                if   s=="ok":  st.markdown(f'<div class="fb-ok"><p class="fb-title">✅ Check-in realizado!</p><p class="fb-nome">{m["nome"]}</p><p class="fb-sub">{m["cargo"]} — {m["localidade"]}</p></div>', unsafe_allow_html=True)
                elif s=="dup": st.markdown(f'<div class="fb-warn"><p class="fb-t">⚠️ Já registrado: {m["nome"]}</p></div>', unsafe_allow_html=True)
                else:          st.markdown('<div class="fb-erro"><p class="fb-t">❌ Código não encontrado</p></div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="fb-idle"><p class="fb-t">QR Code não detectado — tente novamente</p></div>', unsafe_allow_html=True)

    with tab_cod:
        cod_m = st.text_input("Código (ex: CF001)", key="cod_m").upper().strip()
        if st.button("✅ Registrar Presença", type="primary", key="btn_manual"):
            if cod_m:
                m, s = registrar(cod_m, reuniao["id"])
                if   s=="ok":  st.markdown(f'<div class="fb-ok"><p class="fb-title">✅ Check-in realizado!</p><p class="fb-nome">{m["nome"]}</p></div>', unsafe_allow_html=True)
                elif s=="dup": st.markdown(f'<div class="fb-warn"><p class="fb-t">⚠️ Já registrado: {m["nome"]}</p></div>', unsafe_allow_html=True)
                else:          st.markdown('<div class="fb-erro"><p class="fb-t">❌ Código não encontrado</p></div>', unsafe_allow_html=True)

# ===================================================
# RELATORIOS
# ===================================================
elif pagina == "relat":
    st.markdown('''
    <div class="page-header">
        <span class="ph-icon">📊</span>
        <div><p class="ph-title">Relatórios</p>
        <p class="ph-sub">Frequência e presenças por reunião</p></div>
    </div>''', unsafe_allow_html=True)

    reun = load_reunioes(); df_p = load_participantes()
    if not reun: st.info("Nenhuma reunião."); st.stop()

    t_res, t_grf, t_exp = st.tabs(["📋  Presenças", "📈  Gráficos", "⬇️  Exportar"])

    sel    = st.selectbox("📅 Reunião", [r["nome"] for r in reun])
    reuniao = next(r for r in reun if r["nome"] == sel)
    pres   = load_presencas(reuniao["id"])
    ids_p  = {p["id_participante"] for p in pres}
    total  = len(df_p); presente = len(ids_p); ausente = total - presente
    pct    = round(presente/total*100) if total else 0

    with t_res:
        c1,c2,c3 = st.columns(3)
        c1.metric("Total", total)
        c2.metric("Presentes", presente)
        c3.metric("Ausentes",  ausente)
        st.markdown(f'<div class="prog-bar-wrap"><div class="prog-bar-fill" style="width:{pct}%"></div></div><p style="color:#A5A5A5;font-size:.8rem;margin:-14px 0 16px">Frequência: {pct}%</p>', unsafe_allow_html=True)
        if pres:
            df_pr = pd.DataFrame([{
                "Nome": p["participantes"]["nome"],
                "Cargo": p["participantes"]["cargo"],
                "Localidade": p["participantes"]["localidade"],
                "Instrumento": p["participantes"].get("instrumento",""),
                "Hora": p.get("horario","")
            } for p in pres])
            st.dataframe(df_pr, use_container_width=True, hide_index=True)
        else:
            st.info("Nenhuma presença registrada nessa reunião.")

    with t_grf:
        if pres:
            fig1 = px.pie(df_pr, names="Instrumento",
                          title="Presenças por Instrumento",
                          color_discrete_sequence=["#1C3D5A","#49656C","#5DB196","#98CDBD","#A5A5A5","#60b4d4"])
            fig1.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", font_color="#e2e8f0")
            st.plotly_chart(fig1, use_container_width=True)
            fig2 = px.bar(df_pr, x="Cargo", title="Presenças por Cargo",
                          color_discrete_sequence=["#5DB196"])
            fig2.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", font_color="#e2e8f0")
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("Sem dados para gráfico.")

    with t_exp:
        if pres:
            xlsx_buf = BytesIO()
            wb = Workbook(); ws = wb.active; ws.title = "Presenças"
            fill_h = PatternFill("solid", fgColor="1C3D5A")
            font_h = Font(color="FFFFFF", bold=True)
            for ci, h in enumerate(df_pr.columns, 1):
                cell = ws.cell(1,ci,h); cell.fill=fill_h; cell.font=font_h
                cell.alignment=Alignment(horizontal="center")
            for ri, row in df_pr.iterrows():
                for ci,val in enumerate(row,1): ws.cell(ri+2,ci,val)
            for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=22
            wb.save(xlsx_buf); xlsx_buf.seek(0)
            st.download_button("⬇️ Baixar Excel", xlsx_buf,
                               file_name=f"presencas_{sel}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        else:
            st.info("Nenhuma presença para exportar.")

# ===================================================
# MEMBROS
# ===================================================
elif pagina == "membros":
    st.markdown('''
    <div class="page-header">
        <span class="ph-icon">👥</span>
        <div><p class="ph-title">Membros</p>
        <p class="ph-sub">Gerenciar cadastro e gerar QR Code PDF</p></div>
    </div>''', unsafe_allow_html=True)

    INSTRUMENTOS = ["CLARINETE","FLAUTA","ÓRGÃO","SAXOFONE ALTO","SAXOFONE TENOR",
                    "TROMBONE","TROMPETE","TUBA","VIOLINO","VIOLONCELO"]
    CARGOS = ["ENCARREGADO REGIONAL","ENCARREGADO LOCAL",
              "SECRETÁRIO DO GEM","INSTRUTOR","EXAMINADORA DE ORGANISTAS"]

    df = load_participantes()
    t_lista, t_add, t_edit, t_del, t_pdf = st.tabs([
        "📋  Lista",
        "➕  Incluir",
        "✏️  Editar",
        "🗑️  Excluir",
        "🖨️  PDF QR"
    ])

    # --- LISTA ---
    with t_lista:
        st.markdown(f'<div class="sec-label">{len(df)} membros cadastrados</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        fc = c1.selectbox("Cargo",       ["Todos"]+CARGOS,       key="f_cargo")
        fi = c2.selectbox("Instrumento", ["Todos"]+INSTRUMENTOS, key="f_instr")
        dv = df.copy()
        if fc != "Todos": dv = dv[dv["cargo"]==fc]
        if fi != "Todos": dv = dv[dv["instrumento"]==fi]
        cs = [c for c in ["id","nome","cargo","instrumento","localidade"] if c in dv.columns]
        st.dataframe(dv[cs], use_container_width=True, hide_index=True)

    # --- INCLUIR ---
    with t_add:
        st.markdown('<div class="sec-label">Novo Membro</div>', unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        n_id    = col1.text_input("ID (ex: AB030)", key="n_id").upper().strip()
        n_nome  = col2.text_input("Nome completo",  key="n_nome").upper().strip()
        n_cargo = col1.selectbox("Cargo",       CARGOS,       key="n_cargo")
        n_instr = col2.selectbox("Instrumento", INSTRUMENTOS, key="n_instr")
        n_local = st.text_input("Localidade",   key="n_local").upper().strip()
        if st.button("➕ Incluir Membro", type="primary", key="btn_add", use_container_width=True):
            if not n_id or not n_nome or not n_local:
                st.warning("Preencha ID, Nome e Localidade.")
            else:
                chk = supabase.table("participantes").select("id").eq("id",n_id).execute()
                if chk.data: st.error(f"ID {n_id} já existe!")
                else:
                    supabase.table("participantes").insert({"id":n_id,"nome":n_nome,"cargo":n_cargo,"instrumento":n_instr,"localidade":n_local}).execute()
                    st.success(f"✅ {n_nome} incluído com sucesso!"); st.rerun()

    # --- EDITAR ---
    with t_edit:
        if df.empty:
            st.info("Nenhum membro cadastrado.")
        else:
            opts  = df["id"] + " — " + df["nome"]
            sel_e = st.selectbox("Selecione o membro", opts, key="sel_e")
            mid   = sel_e.split(" — ")[0]
            mr    = df[df["id"]==mid].iloc[0]
            col1, col2 = st.columns(2)
            e_nome  = col1.text_input("Nome",       value=mr["nome"],      key="e_nome").upper().strip()
            e_local = col2.text_input("Localidade", value=mr["localidade"], key="e_loc").upper().strip()
            e_cargo = col1.selectbox("Cargo", CARGOS,
                        index=CARGOS.index(mr["cargo"]) if mr["cargo"] in CARGOS else 0, key="e_cargo")
            iv = mr.get("instrumento") or ""
            e_instr = col2.selectbox("Instrumento", INSTRUMENTOS,
                        index=INSTRUMENTOS.index(iv) if iv in INSTRUMENTOS else 0, key="e_instr")
            if st.button("💾 Salvar Alterações", type="primary", key="btn_edit", use_container_width=True):
                supabase.table("participantes").update(
                    {"nome":e_nome,"cargo":e_cargo,"instrumento":e_instr,"localidade":e_local}
                ).eq("id",mid).execute()
                st.success(f"✅ {e_nome} atualizado!"); st.rerun()

    # --- EXCLUIR ---
    with t_del:
        if df.empty:
            st.info("Nenhum membro cadastrado.")
        else:
            opts_d = df["id"] + " — " + df["nome"]
            sel_d  = st.selectbox("Membro para excluir", opts_d, key="sel_d")
            mid_d  = sel_d.split(" — ")[0]
            st.warning(f"⚠️ Excluir **{sel_d}**?\nIsso também remove todas as presenças vinculadas.")
            if st.button("🗑️ Confirmar Exclusão", type="primary", key="btn_del", use_container_width=True):
                supabase.table("presencas").delete().eq("id_participante",mid_d).execute()
                supabase.table("participantes").delete().eq("id",mid_d).execute()
                st.success(f"✅ {mid_d} excluído."); st.rerun()

    # --- PDF QR ---
    with t_pdf:
        if not QR_OK or not RL_OK:
            st.error("❌ Instale: pip install qrcode[pil] reportlab")
        else:
            opcao = st.radio("Selecionar", ["Todos os membros","Membros específicos"], horizontal=True, key="pdf_op")
            if opcao == "Membros específicos":
                opts_p = df["id"] + " — " + df["nome"]
                sels_p = st.multiselect("Selecione", opts_p, key="pdf_sels")
                df_sel = df[df["id"].isin([s.split(" — ")[0] for s in sels_p])]
            else:
                df_sel = df
            st.markdown(f'<div class="sec-label">{len(df_sel)} membro(s) selecionado(s)</div>', unsafe_allow_html=True)
            if st.button("🖨️ Gerar PDF com QR Codes", type="primary", key="btn_pdf", use_container_width=True):
                if df_sel.empty:
                    st.warning("Nenhum membro selecionado.")
                else:
                    with st.spinner("Gerando PDF..."):
                        pdf_b = gerar_pdf_membros(df_sel.to_dict(orient="records"))
                    st.download_button("⬇️ Baixar PDF", data=pdf_b,
                                       file_name="qrcodes_ccb.pdf", mime="application/pdf",
                                       use_container_width=True)
                    st.success(f"✅ {len(df_sel)} cartão(oes) gerado(s)!")
